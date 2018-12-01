#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" auditSecComp_LocalAdminAccess.py

Challenges:
Solutions:
Description:
Attributes:
  __version__ = "1.0.0"
  __project__ = pilot
  __author__ = Jeremy Sung
  __date__ = 6/27/2018 11:31 AM
  __Email__ = Jeremy.Sung@osh.com
Todo:
"""


import os, re
from ldap3 import Server, Connection, SUBTREE, DEREF_ALWAYS
import Enums
import pandas as pd
from openpyxl import Workbook, load_workbook
import bz2
import paramiko
import pprint as pp
import argparse


class auditSecComp_LocalAdminAccess:

  def __init__(self, hostname="", port=22, userid="", passwd="", privateKey="", reportPath=""):
    self.hostname = hostname
    self.port = port if port else 22
    self.userid = userid
    self.passwd = passwd if passwd else ""
    self.privateKey = privateKey

    self.ssh = paramiko.SSHClient()
    self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    if privateKey:
      self.ssh.connect(self.hostname, self.port, username=self.userid, key_filename=self.privateKey)
    else:
      self.ssh.connect(self.hostname, self.port, username=self.userid, password=self.passwd)


    if not os.path.isfile(reportPath):
      self.wb2 = Workbook()
      # self.wb2.remove_sheet(self.wb2.active)
      self.wb2.save(reportPath)

    self.reportPath = reportPath      # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

    self.wb2 = pd.ExcelWriter(self.reportPath, engine='openpyxl')
    # self.wb2.remove_sheet(self.wb2.active)
    self.book = load_workbook(self.reportPath)
    self.wb2.book = self.book

    # Remove the default sheet
    default_sheet = self.wb2.book["Sheet"]
    self.wb2.book.remove(default_sheet)
    # default_sheet = self.wb2.book.get_sheet_by_name("Sheet") # deprecated function get_sheet_by_name (Use wb[sheetname])
    # self.wb2.book.remove_sheet(default_sheet) # deprecated function remove_sheet (Use wb.remove(worksheet) or del wb[sheetname])



  def exeC(self, cmd):

    # ## Ref.  https://stackoverflow.com/questions/10745138/python-paramiko-ssh
    # hostname = '10.1.19.21'  # wms-app91.orchard.osh
    # port = 22
    # userid = 'root'

    # privateKey_root = "C:/admin/id_rsa_root"
    ## ssh.connect(hostname, port, user, passwd)  # Worked!!
    ## ssh.connect(hostname, port, username=user, password=passwd, key_filename="C:/admin/id_rsa_root") # Private Key!!

    # auditSC_LocalAdminAccess = auditSecComp_LocalAdminAccess(hostname=hostname, port=port, userid=userid, privateKey=privateKey_root)

    ## If "sudo" is required the script need to promput user for password though:
    # cmd = "echo {} | sudo -S {}".format(",.SOLpx.01", "cat /etc/sudoers")

    stdin, stdout, stderr = self.ssh.exec_command(cmd)

    result_l = []

    for line in stdout.readlines():
      # if '^#' in line or '^$' in line:
      #   continue
      result_l.append(line)

    # pp.pprint(result_l)

    self.ssh.close()
    return result_l


  def auditAdminAccountsSolo_Helper(self, filename, excludedPhrases):

    entries_l = []

    with open(filename, 'r') as file:

      entries_l.append("Date:\t{}\n".format(Enums.strDateTime))

      for entry in file:
        if any(term in entry for term in excludedPhrases):
          continue
        entries_l.append(entry.strip())

    return entries_l


  def auditAdminAccountsSolo(self, filename):
    """ Expecting: /etc/passwd, /etc/sudoers, or /etc/shadow """

    # hostname = "host_name"
    # CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
    # CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"
    CMD_passwd_shadow = "\nCommand:\t{}".format("date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1")
    CMD_sudoers = "\nCommand:\t{}".format("date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'")

    Excluded_passwd = ('nologin', 'false')
    Excluded_shadow = (':!', ':*')
    Excluded_sudoers = ('^#', '^$')

    resPassword_l = []
    resShadow_l = []
    resSudoers_l = []

    if filename.startswith("passwd"):
      # resPassword_l.append(hostname)
      resPassword_l.append(CMD_passwd_shadow)
      resPassword_l = self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd)
      return resPassword_l

    # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
    if filename.startswith("shadow"):
      # resShadow_l.append(hostname)
      resShadow_l.append(CMD_passwd_shadow)
      resShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))
      return resShadow_l

    if filename.startswith("sudoers"):
      # resSudoers_l.append(hostname)
      resSudoers_l.append(CMD_sudoers)
      resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
      return resSudoers_l

    else:
      print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename))


  def auditAdminAccounts(self, *args):
    """ args represent a tuple or a list of files, such as /etc/passwd, shadow or sudoers """

    if not args:
      print("No file presents. Try again.\n")

    auditResult_l = []
    for filename in args:
      # auditResult_l.extend(self.auditAdminAccountsSolo(filename_sample))

      # hostname = "host_name"
      CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
      CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"

      Excluded_passwd = ('nologin', 'false')
      Excluded_shadow = (':!', ':*')
      Excluded_sudoers = ('^#', '^$')

      resPassword_l = []
      resPasswordShadow_l = []
      resSudoers_l = []

      if filename.startswith("passwd"):

        # resPassword_l.append(hostname)
        resPassword_l.append(CMD_passwd_shadow)
        resPassword_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd))

        # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
        if resPassword_l and filename.startswith("shadow"):
          # resPasswordShadow_l.append(CMD_passwd_shadow)
          resPasswordShadow_l.extend(resPassword_l)
          resPasswordShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))

          # return resPasswordShadow_l
          auditResult_l.extend(resPasswordShadow_l)


        else:
          # return resPassword_l
          auditResult_l.extend(resPassword_l)


      if filename.startswith("sudoers"):
        # resSudoers_l.append(hostname)
        resSudoers_l.append(CMD_sudoers)
        resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
        # return resSudoers_l
        auditResult_l.extend(resSudoers_l)

      # else:
      #   print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename_sample))

    return auditResult_l


  def auditHostAccess_ServerLocalSecureMessages(self):
    """ /etc/secure and /etc/secure.* """

    print(self.DateTime)

    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')

    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            # words =  line.split()
            # reduced_line = " ".join([w for w in words if w not in Excluded])

            # if not any(term in line for term in Excluded):
            if any(term in line for term in Excluded):
              continue

            groups = compiled.search(line)
            # print(groups.group(1).strip())  # May 15 10:16:19
            # print(groups.group(2).strip())  # hostnames
            # print(groups.group(3).strip())  # Messages

            # timestamp, host, message = groups.group(1).strip(), groups.group(2).strip(), groups.group(3)
            # print("{}, {}, {}".format(timestamp, host, message))
            # print("{} -- {} -- {}".format(timestamps, host, messages))

            timestamps_l.append(groups.group(1).strip())
            hostnames_l.append(groups.group(2).strip())
            messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Host Access", index=False)

    self.wb2.save()


  def audit_OracleAccountAccess_DFIODB01_Secure(self):
    """ /etc/secure.* """

    print(self.DateTime)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    # Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')
    Included = ('sudo', 'USER=oracle')
    path = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            groups = compiled.search(line)

            if any(term in groups.group(3) for term in Included):
              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Oracle Account Access", index=False)


    self.wb2.save()



  def auditInformixAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.strDateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "informix"]
    # Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()



            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))

    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Informix Access", index=False)


    self.wb2.save()




  def auditHostRootAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.strDateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "sshd", "root"]
    Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []


    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            # print(line) # b'Jun 17 17:30:01 polling61 /usr/sbin/cron[25801]: (root) CMD ([ -x /usr/lib64/sa/sa1 ] && exec /usr/lib64/sa/sa1 -S ALL 1 1)\n'

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()

            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    # df_secures.to_excel(writer, sheet_name="Root Access", index=False)
    df_secures.to_excel(self.wb2, sheet_name="Root Access", index=False)

    self.wb2.save()






if __name__ == "__main__":
  sol = auditSecComp_LocalAdminAccess()
  



  ###############################################################################################
  ## Argparse Initiatives:
  parser = argparse.ArgumentParser(prog='dev_oshr.py', prefix_chars='-+', description='util to execute Linux commands',
                                   add_help=True, allow_abbrev=True)
  parser.add_argument("-C", "--command", required=True, help="Linux command to execute")
  # parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # parser.add_argument("-C", "--command", action="store_true", help="Linux command to execute")
  ## Positional:
  # parser.add_argument('bar', help='positional bar')
  parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # pp.pprint(parser)

  args = parser.parse_args()
  # pp.pprint(args)

  ## Have to execute with "python" in front!!
  ## C:\staging\python\pilot>python dev_oshr.py --command date
  ## Namespace(command='date', reportPath='C:/staging/python/pilot')

  ## Doesn't work:
  ## C:\staging\python\pilot> dev_oshr.py --command date  # unable to recognize "--command date"


  # print(args.command)
  # print("Command: {}".format(args.command))
  print("Command: {},\tReport: {}".format(args.command, args.report))
  # pp.pprint(args.command)
  # result_l = auditSC_LocalAdminAccess.exeC(args.co)

  # pp.pprint(result_l)


  ##############################################################################################
  ## Worked!
  ## Can run any Linux commands as root@admin01
  ## Get /etc/sudoers as root

  # Ref.  https://stackoverflow.com/questions/10745138/python-paramiko-ssh
  # hostname = '10.1.19.21'  # wms-app91.orchard.osh
  hostname = '10.1.2.210'  # wms-wmosapp01.orchard.osh
  port = 22
  userid = 'root'

  privateKey_root = "C:/admin/id_rsa_root"
  # ssh.connect(hostname, port, user, passwd)  # Worked!!
  # ssh.connect(hostname, port, username=user, password=passwd, key_filename="C:/admin/id_rsa_root") # Private Key!!


  ################# Generic ##################################################################################
  curNum = '00001'
  nextNum = str( int(curNum) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathF = r"C:/staging/python/pilot/testWMS Account Review " + nextNum + r".xlsx"

  auditSC_LocalAdminAccess = auditSecComp_LocalAdminAccess(hostname=hostname, port=port, userid=userid, privateKey=privateKey_root, reportPath=pathF)

  # cmd_sudoers = 'cat /etc/sudoers'
  # cmd_passwd = 'cat /etc/passwd'
  ## If "sudo" is required the script need to promput user for password though:
  # cmd = "echo {} | sudo -S {}".format(",.SOLpx.01", "cat /etc/sudoers")
  # cmd = 'cat /var/log/secure | head -30'
  cmd = "hostname; cat /etc/*release; date"

  # stdin, stdout, stderr = auditSC_LocalAdminAccess.exeC(cmd)

  # result_l = auditSC_LocalAdminAccess.exeC(cmd_passwd)
  result_l = auditSC_LocalAdminAccess.exeC(cmd)



  df_secures = pd.DataFrame({'TimeStamp': result_l})
  df_secures.to_excel(auditSC_LocalAdminAccess.wb2, sheet_name="Host Access", index=False)

  auditSC_LocalAdminAccess.wb2.save()

  # pp.pprint(result_l)


  ################# WMS Servers ####################################################################################


  curWMS = '00001'
  nextWMS = str( int(curWMS) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathWMS = r"C:/staging/python/pilot/testWMS Account Review " + nextWMS + r".xlsx"



  ################# MARS Servers ####################################################################################

  curMARS = '00001'
  nextMARS = str( int(curMARS) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathMARS = r"C:/staging/python/pilot/testMARS Account Review " + nextMARS + r".xlsx"



  ################# Polling Servers ##################################################################################


  curPolling = '00001'
  nextPolling = str( int(curPolling) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathPolling = r"C:/staging/python/pilot/testPolling Account Review " + nextPolling + r".xlsx"


  ############################################################################################################


  # ldapServer = "dc01.orchard.osh"
  # userDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh"
  # passwd = "Ld@p$3cAuth!"
  #
  # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"
  #
  # stg_auditOSHSecurityCompliance_ServerLocalAccess = auditSecComp_ldap(ldapServer, userDN, passwd, reportPath)
  # # stg_auditOSHSecurityCompliance_ServerLocalAccess = auditSecComp_ldap(reportPath)

  # ## /etc/passwd, shadow, sudoers:
  # filename1 = "passwd"
  # filename2 = "shadow"
  # filename3 = "sudoers"
  # result_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.auditAdminAccountsSolo(filename1)
  # filenames = [filename1, filename2, filename3]
  # auditResults_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.auditAdminAccounts(*filenames)
  # for entry in auditResults_l:
  #   print(entry)

  # stg_auditOSHSecurityCompliance_ServerLocalAccess.auditHostAccess_ServerLocalSecureMessages()

  # stg_auditOSHSecurityCompliance_ServerLocalAccess.auditHostAccess_ServerLocalSecureMessages()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.audit_OracleAccountAccess_DFIODB01_Secure()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.auditInformixAccess_Polling_ServerLocalMessages_BZ2()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.auditHostRootAccess_Polling_ServerLocalMessages_BZ2()
