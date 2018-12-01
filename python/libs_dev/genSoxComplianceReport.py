#!/usr/lib/anaconda3/bin/python
#!C:\staging\python\systems\Scripts\python.exe

#  -*- coding: utf-8 -*-
""" genSoxComplianceReport_Ldap.py
Description: Generate PCI/Sox Compliance Audit report - Driver program
Attributes:
  __version__ = "3.0.1"
  __project__ = pilot
  __author__ = Jeremy Sung
  __date__ = 7/31/2018 8:02 AM
  __Email__ = Jeremy.Sung@osh.com
  
Todo:
"""

from ldap3 import Server, Connection, SUBTREE, DEREF_ALWAYS

from auditSoxCompliance import auditSoxCompliance
from auditSoxCompliance_Ldap import auditSoxCompliance_Ldap
from auditSoxCompliance_ServerLocal import auditSoxCompliance_ServerLocal

import pandas as pd
from openpyxl import Workbook, load_workbook

import Utils, Enums
import argparse
import pprint as pp



class genSoxComplianceReport(auditSoxCompliance, auditSoxCompliance_Ldap, auditSoxCompliance_ServerLocal):

  def __init__(self, **kwargs):

    ## Specify ldapServer and authorized credential to perform the audits:
    self.ldapServer = kwargs["ldapServer"]
    self.userDN = kwargs["userDN"]
    self.password = kwargs["password"]

    self.repoReports = kwargs["repoReports"]
    self.projectName = kwargs["projectName"]  # WMS, MARS, Polling

    ## Report Path and repository:
    self.pathCurReport, self.pathNewReport = self.getFilename(projName=self.projectName, dirReports=self.repoReports)
    print(self.pathNewReport)
    ## print(self.pathCurReport)

    if self.projectName == "WMS":
      self.dfLocalAdminAccess_Columns = ['TimeStamp', 'Hosts', 'Messages']
    else:
      self.dfLocalAdminAccess_Columns = ['Dates', 'Hosts', 'Messages']



  def readCurrentSummarySheet(self):
    ##
    ## Add the (Accounts) Summary sheet as the first sheet to the new Workbook:
    ##
    wbOld = pd.ExcelWriter(self.pathCurReport, engine='openpyxl')
    bookOld = load_workbook(self.pathCurReport)
    wbOld.book = bookOld

    summarySheet_str = self.projectName + r' Accounts'
    summarySheet_Name = wbOld.book[summarySheet_str]

    wbNew = pd.ExcelWriter(self.pathNewReport, engine='openpyxl')
    bookNew = load_workbook(self.pathNewReport)
    wbNew.book = bookNew

    bookNew.create_sheet(summarySheet_str, index=0)  ## Create a worksheet at index 0;

    rows = []
    for row in summarySheet_Name.iter_rows():
      row_data = []
      for cell in row:
        row_data.append(cell.value)
      rows.append(row_data)

      bookNew[summarySheet_str].append(row_data)

    wbNew.book.save(self.pathNewReport)



  def genSoxComplianceReport_Ldap(self, **kwargs):

    ## Spawn an LDAP audit instance to carry out Ldap related audits:
    self.auditSC_Ldap = auditSoxCompliance_Ldap(self.ldapServer, self.userDN, self.password, self.pathNewReport)

    self.baseDN_LdapAccessGroups = kwargs["baseDN_LdapAccessGroups"]
    self.searchScope_LdapAccessGroups = kwargs["searchScope_LdapAccessGroups"]

    self.retrieveAttributes_LdapAccessGroups = kwargs["retrieveAttributes_LdapAccessGroups"]
    self.derefAliases_LdapAccessGroups = kwargs["derefAliases_LdapAccessGroups"]

    self.LinuxAccess_l = []

    if self.projectName == "MARS":
      self.searchFilter_App_LdapAccessGroup = kwargs["searchFilter_App_LdapAccessGroup"]
      self.searchFilter_DB_LdapAccessGroup = kwargs["searchFilter_DB_LdapAccessGroup"]
      self.LinuxAccess_l = self.auditSC_Ldap.auditLinuxAdminLdapAccess(self.baseDN_LdapAccessGroups,
                                                             self.searchScope_LdapAccessGroups,
                                                             self.searchFilter_App_LdapAccessGroup,
                                                             self.retrieveAttributes_LdapAccessGroups,
                                                             self.derefAliases_LdapAccessGroups)
      self.LinuxAccess_l.append("\n")
      self.LinuxAccess_l.extend(self.auditSC_Ldap.auditLinuxAdminLdapAccess(self.baseDN_LdapAccessGroups,
                                                             self.searchScope_LdapAccessGroups,
                                                             self.searchFilter_DB_LdapAccessGroup,
                                                             self.retrieveAttributes_LdapAccessGroups,
                                                             self.derefAliases_LdapAccessGroups))

      self.strLDAPAccessTitle = r'LDAP Access (Who has Access to ' + self.projectName + r' Hosts):'
      df_LinuxAccess = pd.DataFrame({self.strLDAPAccessTitle: self.LinuxAccess_l})

      strLDAPAccessSheetName = self.projectName + r" LDAP Access"
      df_LinuxAccess.to_excel(self.auditSC_Ldap.wb2, sheet_name=strLDAPAccessSheetName, index=False)

    else:
      self.searchFilter_LdapAccessGroups = kwargs["searchFilter_LdapAccessGroups"]

      self.LinuxAccess_l = self.auditSC_Ldap.auditLinuxAdminLdapAccess(self.baseDN_LdapAccessGroups,
                                                             self.searchScope_LdapAccessGroups,
                                                             self.searchFilter_LdapAccessGroups,
                                                             self.retrieveAttributes_LdapAccessGroups,
                                                             self.derefAliases_LdapAccessGroups)


      self.strLDAPAccessTitle = r'LDAP Access (Who has Access to ' + self.projectName + r' Hosts):'
      df_LinuxAccess = pd.DataFrame({self.strLDAPAccessTitle : self.LinuxAccess_l})

      strLDAPAccessSheetName = self.projectName + r" LDAP Access"
      df_LinuxAccess.to_excel(self.auditSC_Ldap.wb2, sheet_name=strLDAPAccessSheetName, index=False)

    self.auditSC_Ldap.wb2.save()


    self.baseDN_LinuxGroups = kwargs["baseDN_LinuxGroups"]
    self.searchScope_LinuxGroups = kwargs["searchScope_LinuxGroups"]
    self.gidNumbers_LinuxGroups = kwargs["gidNumbers_LinuxGroups"]
    self.retrieveAttributes_LinuxGroups = kwargs["retrieveAttributes_LinuxGroups"]
    self.derefAliases_LinuxGroups = kwargs["derefAliases_LinuxGroups"]

    gidMembers_Groups_d = self.auditSC_Ldap.listLdapAdminGroupMembership(self.baseDN_LinuxGroups,
                                                                    self.searchScope_LinuxGroups,
                                                                    self.retrieveAttributes_LinuxGroups,
                                                                    self.gidNumbers_LinuxGroups)

    LdapAdminGroups_l = []
    for grpID in gidMembers_Groups_d.keys():

      LdapAdminGroups_l.append("\n")
      LdapAdminGroups_l.append('Linux Group ID [{}]:'.format(grpID))
      LdapAdminGroups_l.append(
        r'date && ldapsearch -H ldaps://' + self.ldapServer + r':636 -x -D CN=<ldap_user>,OU=Users,OU=Administrative,DC=Orchard,DC=osh" -w "<ldap_passwd>" -b "' + self.baseDN_LinuxGroups +
        r'" -s sub "(gidNumber=' + grpID + r')" ' + " ".join(self.retrieveAttributes_LinuxGroups) +
        r' | grep "^dn\|^gid"')

      LdapAdminGroups_l.append("\n")
      LdapAdminGroups_l.append('{}'.format(self.auditSC_Ldap.DateTime))

      countGroupMembers = 0
      for member in gidMembers_Groups_d[grpID]:
        LdapAdminGroups_l.append(member)
        countGroupMembers += 1

      LdapAdminGroups_l.append('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
      LdapAdminGroups_l.append("\n")

    self.strSheetName_LDAP_Groups = self.projectName + " LDAP Groups"
    df_wmsLdapAdminGroups = pd.DataFrame({'LADP Groups Access (Who is assigned to what group):': LdapAdminGroups_l})
    df_wmsLdapAdminGroups.to_excel(self.auditSC_Ldap.wb2, sheet_name=self.strSheetName_LDAP_Groups, index=False)

    self.auditSC_Ldap.wb2.save()


  def genSoxComplianceReport_ServerLocal(self, **kwargs):
    ## Audit Linux Server Local - Accounts & Logs: /etc/passwd, shadow, sudoers; /var/logs/secure, messages, etc.
    ## /etc/sudoers

    self.port = kwargs["port"]
    self.userid = kwargs["userid"]
    self.privateKey = kwargs["privateKey"]  ## privateKey_root

    strServers = r"auditSC_Servers_" + self.projectName
    self.prjServers = getattr(Enums, strServers)

    ##
    ## Audit Local Admin Accounts: /etc/passwd, shadow, and sudoers:
    ##
    resLocalAdminAccounts_Servers_d = {}
    resLocalAdminAccounts_Servers_l = []
    resLocalAdminAccess_Servers_d = {}

    df_AdminAccess_Servers_Frames_l = []
    df_OracleAccess_Servers_Frames_l = []

    resOracleAccess_Servers_d = {}

    ## Polling:
    resLocalPasswdShadow_Servers_l = []
    resSudoers_Servers_l = []
    resInformixAccess_Servers_d = {}
    resRootAccess_Servers_d = {}

    df_InformixAccess_Servers_Frames_l = []
    df_RootAccess_Servers_Frames_l = []

    if self.projectName == "Polling":

      for server in self.prjServers:
        self.auditSC_LocalAccess_Server = auditSoxCompliance_ServerLocal(hostname=server,
                                                                    port=self.port,
                                                                    userid=self.userid,
                                                                    privateKey=self.privateKey,
                                                                    reportPath=self.pathNewReport)

        ##
        ## Audit Linux Admin Accounts on Server Local:
        ##
        ## Result of /etc/password & shadow audit:
        resPasswordShadow_RemoteServer_l = self.auditSC_LocalAccess_Server.scanPasswdShadow_RemoteServer()
        ## Result of /etc/sudoers audit:
        resSudoers_RemoteServer_l = self.auditSC_LocalAccess_Server.scanSudoers_RemoteServer()

        resLocalPasswdShadow_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_passwd_shadow)
        resLocalPasswdShadow_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resLocalPasswdShadow_Servers_l.extend(resPasswordShadow_RemoteServer_l)
        resLocalPasswdShadow_Servers_l.append("\n")

        resSudoers_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_sudoers)
        resSudoers_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resSudoers_Servers_l.extend(resSudoers_RemoteServer_l)
        resSudoers_Servers_l.append("\n")

        ## Result per Server:
        resLocalAdminAccounts_Servers_d[server] = [resPasswordShadow_RemoteServer_l, resSudoers_RemoteServer_l]

        ##
        ## Audit Informix & Root Access to Polling Server Local:
        ##
        filterInformixAccess = {"Include_all": ["sudo", "informix", ], "Include_any": [], "Exclude": []}
        timestamps_server_informix_l, hostnames_server_informix_l, messages_server_informix_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogMessages_BZ2_Polling(
          **filterInformixAccess)

        resInformixAccess_Server_d = {}  ## Key move!!! Must initialize (reset) this Dict here, instead of outside the loop!

        resInformixAccess_Server_d["Dates"] = timestamps_server_informix_l
        resInformixAccess_Server_d["Hosts"] = hostnames_server_informix_l
        resInformixAccess_Server_d["Messages"] = messages_server_informix_l

        resInformixAccess_Servers_d[server] = resInformixAccess_Server_d

        df_InformixAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resInformixAccess_Server_d,
                                                                                    columns=self.dfLocalAdminAccess_Columns))
        resRootAccess_Server_d = {}  ## Key Move!!

        filterRootAccess = {"Include_all_A": ["sudo", "root", ], "Include_all_B": ["sshd", "root", ],
                            "Include_any": ["sudo", "sshd", ], "Exclude": ["nagios", ]}
        timestamps_server_rootaccess_l, hostnames_server_rootaccess_l, messages_server_rootaccess_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogMessages_BZ2_Root_Polling(
          **filterRootAccess)

        resRootAccess_Server_d["Dates"] = timestamps_server_rootaccess_l
        resRootAccess_Server_d["Hosts"] = hostnames_server_rootaccess_l
        resRootAccess_Server_d["Messages"] = messages_server_rootaccess_l

        resRootAccess_Servers_d[server] = resRootAccess_Server_d
        df_RootAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resRootAccess_Server_d,
                                                                        columns=self.dfLocalAdminAccess_Columns))

      self.strSheetName_Local_Access = self.projectName + " Local Access"
      df_LocalAdminAccounts_Servers = pd.DataFrame({"Local Access" : resLocalPasswdShadow_Servers_l})
      df_LocalAdminAccounts_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Local_Access, index=False)

      self.strSheetName_Sudo_File = self.projectName + " SUDO File"
      df_LocalAdminAccounts_Servers = pd.DataFrame( {"SUDO File" : resSudoers_Servers_l} )
      df_LocalAdminAccounts_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Sudo_File, index=False)

      ##
      ## For Audit: Form DataFrame - Informix & Root Access to Polling Server Local:
      ##
      # pp.pprint(df_InformixAccess_Servers_Frames_Polling_l)
      self.strSheetName_Informix_Access = self.projectName + " Informix Access"
      df_InformixAccess_Servers = pd.concat( df_InformixAccess_Servers_Frames_l )
      df_InformixAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Informix_Access, index=False)

      self.strSheetName_Root_Access = self.projectName + " Root Access"
      df_RootAccess_Servers = pd.concat(df_RootAccess_Servers_Frames_l)
      df_RootAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Root_Access, index=False)

      ##
      ## Wrap Up and Save updates to Excel Workbook:
      ##
      self.auditSC_LocalAccess_Server.wb2.save()

      ########################################################################################################
      ##
      ## Save updates on Excel Workbook:
      ##
      ## self.auditSC_LocalAccess_Server.wb2.save()
      ##
      ## Remove the redundant default worksheet, titled "Sheet"
      ##
      default_sheet = self.auditSC_LocalAccess_Server.wb2.book["Sheet"]
      self.auditSC_LocalAccess_Server.wb2.book.remove(default_sheet)
      self.auditSC_LocalAccess_Server.wb2.save()

    elif self.projectName == "WMS":

      for server in self.prjServers:
        self.auditSC_LocalAccess_Server = auditSoxCompliance_ServerLocal(hostname=server,
                                                                    port=self.port,
                                                                    userid=self.userid,
                                                                    privateKey=self.privateKey,
                                                                    reportPath=self.pathNewReport)
        ##
        ## Audit Linux Admin Accounts on Server Local:
        ##
        ## Result of /etc/password & shadow audit:
        resPasswordShadow_RemoteServer_l = self.auditSC_LocalAccess_Server.scanPasswdShadow_RemoteServer()
        ## Result of /etc/sudoers audit:
        resSudoers_RemoteServer_l = self.auditSC_LocalAccess_Server.scanSudoers_RemoteServer()

        resLocalAdminAccounts_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_passwd_shadow)
        resLocalAdminAccounts_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resLocalAdminAccounts_Servers_l.extend(resPasswordShadow_RemoteServer_l)
        resLocalAdminAccounts_Servers_l.append("\n")

        resLocalAdminAccounts_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_sudoers)
        resLocalAdminAccounts_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resLocalAdminAccounts_Servers_l.extend(resSudoers_RemoteServer_l)
        resLocalAdminAccounts_Servers_l.append("\n")

        ## Result per Server:
        resLocalAdminAccounts_Servers_d[server] = [resPasswordShadow_RemoteServer_l, resSudoers_RemoteServer_l]

        ##
        ## Audit Admin "Host Access" to Server Local - Host Access:
        ##
        ## if self.projectName == "WMS":
        filterHostAccess = {"Include_All": [], "Include_Any": [],
                            "Exclude": ["rfuser", "disconnect", "pam_unix", "subsystem request for sftp"]}
        timestamps_server_l, hostnames_server_l, messages_server_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogSecure_Generic(
          **filterHostAccess)

        resLocalAdminAccess_Server_d = {}  ## Key Move!!

        resLocalAdminAccess_Server_d["TimeStamp"] = timestamps_server_l
        resLocalAdminAccess_Server_d["Hosts"] = hostnames_server_l
        resLocalAdminAccess_Server_d["Messages"] = messages_server_l

        resLocalAdminAccess_Servers_d[server] = resLocalAdminAccess_Server_d

        df_AdminAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resLocalAdminAccess_Server_d,
                                                                         columns=self.dfLocalAdminAccess_Columns))

        ##
        ## Audit "Oracle Account Access" to Server Local:
        ##
        filterOracleAccess = {"Include_All": ["sudo", "USER=oracle"], "Include_Any": [], "Exclude": []}

        timestamps_server_oracle_l, hostnames_server_oracle_l, messages_server_oracle_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogSecures(
          **filterOracleAccess)

        resOracleAccess_Server_d = {}  ## Key move!!! Must initialize (reset) this Dict here, instead of outside the loop!

        resOracleAccess_Server_d["TimeStamp"] = timestamps_server_oracle_l
        resOracleAccess_Server_d["Hosts"] = hostnames_server_oracle_l
        resOracleAccess_Server_d["Messages"] = messages_server_oracle_l

        resOracleAccess_Servers_d[server] = resOracleAccess_Server_d  # Not used at all!!

        df_OracleAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resOracleAccess_Server_d,
                                                                               columns=self.dfLocalAdminAccess_Columns))


      ##
      ## For Audit: Local Admin Accounts: /etc/passwd, shadow, and sudoers:
      ##
      self.strSheetName_Local_Sudo_Access = self.projectName + " Local Access - SUDO"
      df_LocalAdminAccounts_Servers = pd.DataFrame({"Local Access and SUDO File": resLocalAdminAccounts_Servers_l})
      df_LocalAdminAccounts_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Local_Sudo_Access,
                                             index=False)
      ##
      ## For Audit: Admin Access to Server Local:
      ##
      self.strSheetName_Host_Access = self.projectName + " Host Access"
      df_AdminAccess_Servers = pd.concat(df_AdminAccess_Servers_Frames_l)
      df_AdminAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Host_Access, index=False)

      ## For Audit: Oracle Account Access
      self.strSheetName_Oracle_Access = self.projectName + " Oracle Account Access"
      df_OracleAccess_Servers = pd.concat(df_OracleAccess_Servers_Frames_l)
      df_OracleAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Oracle_Access,
                                            index=False)



      ########################################################################################################
      ##
      ## Save updates on Excel Workbook:
      ##
      self.auditSC_LocalAccess_Server.wb2.save()
      ##
      ## Remove the redundant default worksheet, titled "Sheet"
      ##
      default_sheet = self.auditSC_LocalAccess_Server.wb2.book["Sheet"]
      self.auditSC_LocalAccess_Server.wb2.book.remove(default_sheet)
      self.auditSC_LocalAccess_Server.wb2.save()


    elif self.projectName == "MARS":

      for server in self.prjServers:
        self.auditSC_LocalAccess_Server = auditSoxCompliance_ServerLocal(hostname=server,
                                                                         port=self.port,
                                                                         userid=self.userid,
                                                                         privateKey=self.privateKey,
                                                                         reportPath=self.pathNewReport)
        ##
        ## Audit Linux Admin Accounts on Server Local:
        ##
        ## Result of /etc/password & shadow audit:
        resPasswordShadow_RemoteServer_l = self.auditSC_LocalAccess_Server.scanPasswdShadow_RemoteServer()
        ## Result of /etc/sudoers audit:
        resSudoers_RemoteServer_l = self.auditSC_LocalAccess_Server.scanSudoers_RemoteServer()

        resLocalAdminAccounts_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_passwd_shadow)
        resLocalAdminAccounts_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resLocalAdminAccounts_Servers_l.extend(resPasswordShadow_RemoteServer_l)
        resLocalAdminAccounts_Servers_l.append("\n")

        resLocalAdminAccounts_Servers_l.append(server + ":~ # " + self.auditSC_LocalAccess_Server.CMD_sudoers)
        resLocalAdminAccounts_Servers_l.append(self.auditSC_LocalAccess_Server.DateTime)
        resLocalAdminAccounts_Servers_l.extend(resSudoers_RemoteServer_l)
        resLocalAdminAccounts_Servers_l.append("\n")

        ## Result per Server:
        resLocalAdminAccounts_Servers_d[server] = [resPasswordShadow_RemoteServer_l, resSudoers_RemoteServer_l]  # Not used.

        ##
        ## Audit "Oracle Account Access" to Server Local:
        ##
        filterOracleAccess = {"Include_All": ["sudo", "USER=oracle"], "Include_Any": [], "Exclude": []}

        timestamps_server_oracle_l, hostnames_server_oracle_l, messages_server_oracle_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogSecures(
          **filterOracleAccess)

        resOracleAccess_Server_d = {}  ## Key move!!! Must initialize (reset) this Dict here, instead of outside the loop!

        resOracleAccess_Server_d["Dates"] = timestamps_server_oracle_l
        resOracleAccess_Server_d["Hosts"] = hostnames_server_oracle_l
        resOracleAccess_Server_d["Messages"] = messages_server_oracle_l

        resOracleAccess_Servers_d[server] = resOracleAccess_Server_d  # Not used at all!!

        df_OracleAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resOracleAccess_Server_d,
                                                                               columns=self.dfLocalAdminAccess_Columns))

        ##
        ## Audit Admin "Host Access" to Server Local - Host Access:
        ##
        filterHostAccess = {"Include_All": [], "Include_Any": [],
                            "Exclude": ['rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp',
                                        'marssftp from 10.1.1.63', ]}

        timestamps_server_l, hostnames_server_l, messages_server_l = self.auditSC_LocalAccess_Server.scanHostAccess_RemoteServer_VarLogSecure_Generic(
          **filterHostAccess)

        resLocalAdminAccess_Server_d = {}  ## Key Move!!

        resLocalAdminAccess_Server_d["Dates"] = timestamps_server_l
        resLocalAdminAccess_Server_d["Hosts"] = hostnames_server_l
        resLocalAdminAccess_Server_d["Messages"] = messages_server_l

        resLocalAdminAccess_Servers_d[server] = resLocalAdminAccess_Server_d   # Not used at all?!

        df_AdminAccess_Servers_Frames_l.append(pd.DataFrame.from_records(resLocalAdminAccess_Server_d,
                                                                         columns=self.dfLocalAdminAccess_Columns))


      ##
      ## For Audit: Local Admin / Sudo Accounts: /etc/passwd, shadow, and sudoers:
      ##
      self.strSheetName_Local_Sudo_Access = self.projectName + " Local Access - SUDO"
      df_LocalAdminAccounts_Servers = pd.DataFrame({"Local Access and SUDO File": resLocalAdminAccounts_Servers_l})
      df_LocalAdminAccounts_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Local_Sudo_Access,
                                             index=False)

      ## For Audit: Oracle Account Access
      self.strSheetName_Oracle_Access = self.projectName + " Oracle Account Access"
      df_OracleAccess_Servers = pd.concat(df_OracleAccess_Servers_Frames_l)
      df_OracleAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name=self.strSheetName_Oracle_Access,
                                            index=False)

      ##
      ## For Audit: Admin Host Access to Server Local:
      ##
      df_AdminAccess_Servers = pd.concat(df_AdminAccess_Servers_Frames_l)
      df_AdminAccess_Servers.to_excel(self.auditSC_LocalAccess_Server.wb2, sheet_name="Host Access", index=False)

      ########################################################################################################
      ##
      ## Save updates on Excel Workbook:
      ##
      self.auditSC_LocalAccess_Server.wb2.save()
      ##
      ## Remove the redundant default worksheet, titled "Sheet"
      ##
      default_sheet = self.auditSC_LocalAccess_Server.wb2.book["Sheet"]
      self.auditSC_LocalAccess_Server.wb2.book.remove(default_sheet)
      self.auditSC_LocalAccess_Server.wb2.save()



if __name__ == "__main__":
  ##
  ##  WMS :
  ##
  ## Dev:
  # kwargs_wms = {"ldapServer" : "dc01.orchard.osh",
  #               "userDN" : "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
  #               "password" : "Ld@p$3cAuth!",
  #
  #               "repoReports" : "C:/staging/python/pilot",
  #               "baseDN_LdapAccessGroups" : "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
  #               "searchScope_LdapAccessGroups" : SUBTREE,
  #
  #               "projectName" : "WMS",
  #               "searchFilter_LdapAccessGroups" : '(cn=Linux-WMS-Access*)',
  #
  #               "retrieveAttributes_LdapAccessGroups" : ['cn', 'member'],
  #               "derefAliases_LdapAccessGroups" : DEREF_ALWAYS,
  #               "baseDN_LinuxGroups" : "dc=orchard,dc=osh",
  #               "searchScope_LinuxGroups" : SUBTREE,
  #               "gidNumbers_LinuxGroups" : ['10000', '10001', '10004', '10005'],
  #               "retrieveAttributes_LinuxGroups" : ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
  #               "derefAliases_LinuxGroups" : DEREF_ALWAYS,
  #
  #               "port" : 22,
  #               "userid" : "root",
  #               "privateKey" : "c:/admin/id_rsa_root",
  #               }

  ## Prod at admin01:
  kwargs_wms = {"ldapServer" : "dc01.orchard.osh",
                "userDN" : "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
                "password" : "Ld@p$3cAuth!",

                "repoReports" : "/root/data/auditReports/",
                "baseDN_LdapAccessGroups" : "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
                "searchScope_LdapAccessGroups" : SUBTREE,

                "projectName" : "WMS",
                "searchFilter_LdapAccessGroups" : '(cn=Linux-WMS-Access*)',

                "retrieveAttributes_LdapAccessGroups" : ['cn', 'member'],
                "derefAliases_LdapAccessGroups" : DEREF_ALWAYS,
                "baseDN_LinuxGroups" : "dc=orchard,dc=osh",
                "searchScope_LinuxGroups" : SUBTREE,
                "gidNumbers_LinuxGroups" : ['10000', '10001', '10004', '10005'],
                "retrieveAttributes_LinuxGroups" : ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
                "derefAliases_LinuxGroups" : DEREF_ALWAYS,

                "port" : 22,
                "userid" : "root",
                "privateKey" : "/root/.ssh/id_rsa",
                }

  generateSoxComplianceReport_WMS = genSoxComplianceReport(**kwargs_wms)
  generateSoxComplianceReport_WMS.genSoxComplianceReport_Ldap(**kwargs_wms)
  generateSoxComplianceReport_WMS.genSoxComplianceReport_ServerLocal(**kwargs_wms)
  generateSoxComplianceReport_WMS.readCurrentSummarySheet()

  ###################################################################################################################
  ##
  ##  MARS :
  ##
  ## Dev:
  # kwargs_mars = {"ldapServer" : "dc01.orchard.osh",
  #               "userDN" : "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
  #               "password" : "Ld@p$3cAuth!",
  #
  #               "repoReports" : "C:/staging/python/pilot",
  #               "baseDN_LdapAccessGroups" : "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
  #               "searchScope_LdapAccessGroups" : SUBTREE,
  #
  #               "projectName" : "MARS",
  #               "searchFilter_App_LdapAccessGroup" : '(cn=Linux-MARS-App-Access*)',
  #               "searchFilter_DB_LdapAccessGroup" : '(cn=Linux-MARS-DB-Access*)',
  #               "retrieveAttributes_LdapAccessGroups": ['cn', 'member'],
  #               "derefAliases_LdapAccessGroups": DEREF_ALWAYS,
  #
  #               "baseDN_LinuxGroups": "dc=orchard,dc=osh",
  #               "searchScope_LinuxGroups": SUBTREE,
  #               "gidNumbers_LinuxGroups": ['10000', '10001', '10004', '10005'],
  #               "retrieveAttributes_LinuxGroups": ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
  #               "derefAliases_LinuxGroups": DEREF_ALWAYS,
  #
  #               "port": 22,
  #               "userid": "root",
  #               "privateKey": "c:/admin/id_rsa_root",
  #               }

  ## Prod at admin01:
  kwargs_mars = {"ldapServer" : "dc01.orchard.osh",
                "userDN" : "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
                "password" : "Ld@p$3cAuth!",

                "repoReports" : "/root/data/auditReports/",
                "baseDN_LdapAccessGroups" : "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
                "searchScope_LdapAccessGroups" : SUBTREE,

                "projectName" : "MARS",
                "searchFilter_App_LdapAccessGroup" : '(cn=Linux-MARS-App-Access*)',
                "searchFilter_DB_LdapAccessGroup" : '(cn=Linux-MARS-DB-Access*)',
                "retrieveAttributes_LdapAccessGroups": ['cn', 'member'],
                "derefAliases_LdapAccessGroups": DEREF_ALWAYS,

                "baseDN_LinuxGroups": "dc=orchard,dc=osh",
                "searchScope_LinuxGroups": SUBTREE,
                "gidNumbers_LinuxGroups": ['10000', '10001', '10004', '10005'],
                "retrieveAttributes_LinuxGroups": ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
                "derefAliases_LinuxGroups": DEREF_ALWAYS,

                "port": 22,
                "userid": "root",
                "privateKey": "/root/.ssh/id_rsa"
                }

  generateSoxComplianceReport_MARS = genSoxComplianceReport(**kwargs_mars)
  generateSoxComplianceReport_MARS.genSoxComplianceReport_Ldap(**kwargs_mars)
  generateSoxComplianceReport_MARS.genSoxComplianceReport_ServerLocal(**kwargs_mars)
  generateSoxComplianceReport_MARS.readCurrentSummarySheet()


  ##################################################################################################################
  ##
  ##  Polling :
  ##
  ## Dev:
  # kwargs_polling = {"ldapServer": "dc01.orchard.osh",
  #               "userDN": "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
  #               "password": "Ld@p$3cAuth!",
  #
  #               "repoReports": "C:/staging/python/pilot",
  #               "baseDN_LdapAccessGroups": "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
  #               "searchScope_LdapAccessGroups": SUBTREE,
  #
  #               "projectName": "Polling",
  #               "searchFilter_LdapAccessGroups": '(cn=Linux-Polling-Access*)',
  #
  #               "retrieveAttributes_LdapAccessGroups": ['cn', 'member'],
  #               "derefAliases_LdapAccessGroups": DEREF_ALWAYS,
  #               "baseDN_LinuxGroups": "dc=orchard,dc=osh",
  #               "searchScope_LinuxGroups": SUBTREE,
  #               "gidNumbers_LinuxGroups": ['10000', '10001', '10003'],
  #               "retrieveAttributes_LinuxGroups": ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
  #               "derefAliases_LinuxGroups": DEREF_ALWAYS,
  #
  #               "port": 22,
  #               "userid": "root",
  #               "privateKey": "c:/admin/id_rsa_root",
  #               }

  ### Prod at admin01:
  kwargs_polling = {"ldapServer": "dc01.orchard.osh",
                "userDN": "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh",
                "password": "Ld@p$3cAuth!",

                "repoReports": "/root/data/auditReports/",
                "baseDN_LdapAccessGroups": "ou=Groups,ou=Administrative,dc=orchard,dc=osh",
                "searchScope_LdapAccessGroups": SUBTREE,

                "projectName": "Polling",
                "searchFilter_LdapAccessGroups": '(cn=Linux-Polling-Access*)',

                "retrieveAttributes_LdapAccessGroups": ['cn', 'member'],
                "derefAliases_LdapAccessGroups": DEREF_ALWAYS,
                "baseDN_LinuxGroups": "dc=orchard,dc=osh",
                "searchScope_LinuxGroups": SUBTREE,
                "gidNumbers_LinuxGroups": ['10000', '10001', '10003'],
                "retrieveAttributes_LinuxGroups": ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory'],
                "derefAliases_LinuxGroups": DEREF_ALWAYS,

                "port": 22,
                "userid": "root",
                "privateKey": "/root/.ssh/id_rsa",
                }

  generateSoxComplianceReport_Polling = genSoxComplianceReport(**kwargs_polling)
  generateSoxComplianceReport_Polling.genSoxComplianceReport_Ldap(**kwargs_polling)
  generateSoxComplianceReport_Polling.genSoxComplianceReport_ServerLocal(**kwargs_polling)
  generateSoxComplianceReport_Polling.readCurrentSummarySheet()


  # ###############################################################################################
  #
  # ## Argparse Initiatives:
  # ## Allow interaction with console:
  #
  # parser = argparse.ArgumentParser(prog='dev_oshr.py', prefix_chars='-+', description='util to execute Linux commands',
  #                                  add_help=True, allow_abbrev=True)
  # parser.add_argument("-C", "--command", required=True, help="Linux command to execute")
  # # parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # # parser.add_argument("-C", "--command", action="store_true", help="Linux command to execute")
  # ## Positional:
  # # parser.add_argument('bar', help='positional bar')
  # parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # # pp.pprint(parser)
  #
  # args = parser.parse_args()
  # # pp.pprint(args)
  #
  # ## Have to execute with "python" in front!!
  # ## C:\staging\python\pilot>python dev_oshr.py --command date
  # ## Namespace(command='date', reportPath='C:/staging/python/pilot')
  #
  # ## Doesn't work:
  # ## C:\staging\python\pilot> dev_oshr.py --command date  # unable to recognize "--command date"
  #
  #
  # # print(args.command)
  # # print("Command: {}".format(args.command))
  # print("Command: {},\tReport: {}".format(args.command, args.report))
  # # pp.pprint(args.command)
  # # result_l = auditSC_LocalAdminAccess.exeC(args.co)
  #
  # # pp.pprint(result_l)