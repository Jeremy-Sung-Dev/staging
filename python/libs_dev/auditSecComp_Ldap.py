#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" auditSecComp_Ldap.py

Challenges:
Solutions:
Description:
Attributes:
  __version__ = "1.0.0"
  __project__ = pilot
  __author__ = Jeremy Sung
  __date__ = 6/27/2018 11:31 AM
  __Email__ = Jeremy.Sung@osh.com
Todo:
"""

import os, re
from ldap3 import Server, Connection, SUBTREE, DEREF_ALWAYS
import Enums
import pandas as pd
from openpyxl import Workbook, load_workbook
import bz2
import paramiko
import pprint as pp
import argparse




class auditSecComp_ldap:

  def __init__(self, ldapServer, userDN, password, reportPath):
    Domain = '.orchard.osh'
    if '.' not in ldapServer:
      self.ldapServer = ldapServer + Domain

    self.ldapServer = "dc01.orchard.osh" if not ldapServer else ldapServer
    self.UserDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh" if not userDN else userDN
    self.Password = "Ld@p$3cAuth!" if not password else password
    ## self.Port = 636

    ## first open a connection to the server:
    server = Server(self.ldapServer)
    self.conn = Connection(server, user=self.UserDN, password=self.Password)

    self.lastMonth = Enums.strLastMonth  # May
    self.DateTime = Enums.strDateTime

    if not os.path.isfile(reportPath):
      self.wb2 = Workbook()
      # self.wb2.remove_sheet(self.wb2.active)
      self.wb2.save(reportPath)

    self.reportPath = reportPath      # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

    self.wb2 = pd.ExcelWriter(self.reportPath, engine='openpyxl')
    # self.wb2.remove_sheet(self.wb2.active)
    self.book = load_workbook(self.reportPath)
    self.wb2.book = self.book

    # Remove the default sheet
    default_sheet = self.wb2.book["Sheet"]
    self.wb2.book.remove(default_sheet)
    # default_sheet = self.wb2.book.get_sheet_by_name("Sheet") # deprecated function get_sheet_by_name (Use wb[sheetname])
    # self.wb2.book.remove_sheet(default_sheet) # deprecated function remove_sheet (Use wb.remove(worksheet) or del wb[sheetname])


  def getLDAPLinuxAccessGroupMembers(self, baseDN, searchScope, searchFilter, retrieveAttributes, derefAliases):

    self.conn.open()
    self.conn.bind()

    wmsLinuxAccessGroupMembers_l = []

    derefAliases = DEREF_ALWAYS if not derefAliases else derefAliases

    ## Review WMS Access Groups - Print group members for server access.
    self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter, attributes=retrieveAttributes,
                dereference_aliases=derefAliases)

    ## WMS LDAP access:
    responses_d = self.conn.response[0]
    attributes_d = responses_d['attributes']

    for member in attributes_d['member']:
      wmsLinuxAccessGroupMembers_l.append(member)

    self.conn.unbind()
    return wmsLinuxAccessGroupMembers_l


  def getLinuxGroupMembers(self, baseDN, searchScope, retrieveAttributes, gidNumbers):

    self.conn.open()
    self.conn.bind()

    gidMembers_d = {}
    for gidNum in gidNumbers:
      searchFilter = '(gidNumber=' + gidNum + ')'

      self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter,
                  attributes=retrieveAttributes)

      gidMembers_l = []
      for member_d in self.conn.response:
        for key, value in member_d.items():
          if key == 'dn':
            gidMembers_l.append(member_d[key])
          else:
            continue

      gidMembers_d[gidNum] = gidMembers_l

    self.conn.unbind()
    return gidMembers_d


  ## WMS - Linux Audit Hosts - Verify local Server Accounts and Sudo file:
  # auditHosts = ['wms-wmosapp01', 'wms-wmosdb01', 'wms-sciapp01', 'wms-scidb01', 'wms-ifeeapp01']







if __name__ == "__main__":


  # sol = auditSecComp_ldap()

  ldapServer = "dc01.orchard.osh"
  userDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh"
  password = "Ld@p$3cAuth!"

  #############################################################################################################


  curWMS = '00001'
  nextWMS = str( int(curWMS) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathWMS = r"C:/staging/python/pilot/testWMS Account Review " + nextWMS + r".xlsx"

  stg_auditOSHSecurityComplianceWMS = auditSecComp_ldap(ldapServer, userDN, password, pathWMS)

  ## WMS - Access Groups:
  baseDN_WMS_AccessGroup = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  searchScope_WMS_AccessGroup = SUBTREE
  searchFilter_WMS_AccessGroup = '(cn=Linux-WMS-Access*)'
  retrieveAttributes_WMS_AccessGroup = ['cn', 'member']
  derefAliases_WMS_AccessGroup = DEREF_ALWAYS

  wmsLinuxAccessGroupMembers_l = stg_auditOSHSecurityComplianceWMS.getLDAPLinuxAccessGroupMembers(baseDN_WMS_AccessGroup, searchScope_WMS_AccessGroup,
                                                                                                  searchFilter_WMS_AccessGroup,
                                                                                                  retrieveAttributes_WMS_AccessGroup, derefAliases_WMS_AccessGroup)
  # pp.pprint(wmsLinuxAccessGroupMembers_l)

  df_secures = pd.DataFrame({'WMS': wmsLinuxAccessGroupMembers_l})
  df_secures.to_excel(stg_auditOSHSecurityComplianceWMS.wb2, sheet_name="WMS LDAP Access", index=False)

  stg_auditOSHSecurityComplianceWMS.wb2.save()

  ## WMS - Linux groups:
  baseDN_WMS_LinuxGroups = "dc=orchard,dc=osh"
  searchScope_WMS_LinuxGroups = SUBTREE
  gidNumbers_WMS_LinuxGroups = ['10000', '10001', '10004', '10005']
  retrieveAttributes_WMS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  derefAliases_WMS_LinuxGroups = DEREF_ALWAYS

  gidMembers_d = stg_auditOSHSecurityComplianceWMS.getLinuxGroupMembers(baseDN_WMS_LinuxGroups, searchScope_WMS_LinuxGroups,
                                                                        retrieveAttributes_WMS_LinuxGroups, gidNumbers_WMS_LinuxGroups)
  WMS_Members_l = []
  for grpID in gidMembers_d.keys():

    # print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceWMS.DateTime, grpID))
    WMS_Members_l.append('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceWMS.DateTime, grpID))

    countGroupMembers = 0
    for member in gidMembers_d[grpID]:
      # print(member)
      WMS_Members_l.append(member)
      countGroupMembers += 1

    # print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
    WMS_Members_l.append('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))

  df_secures = pd.DataFrame({'WMS': WMS_Members_l})
  df_secures.to_excel(stg_auditOSHSecurityComplianceWMS.wb2, sheet_name="WMS LDAP Groups", index=False)

  stg_auditOSHSecurityComplianceWMS.wb2.save()


  #############################################################################################################


  #############################################################################################################
  ## MARS - Access Groups:
  baseDN_MARS_AccessGroup = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  searchScope_MARS_AccessGroup = SUBTREE
  searchFilter_App_MARS_AccessGroup = '(cn=Linux-MARS-App-Access*)'
  searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  retrieveAttributes_MARS_AccessGroup = ['cn', 'member']
  derefAliases_MARS_AccessGroup = DEREF_ALWAYS


  curMARS = '00001'
  nextMARS = str( int(curMARS) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathMARS = r"C:/staging/python/pilot/testMARS Account Review " + nextMARS + r".xlsx"

  stg_auditOSHSecurityComplianceMARS = auditSecComp_ldap(ldapServer, userDN, password, pathMARS)

  marsLinuxAccessGroupMembers_l = stg_auditOSHSecurityComplianceMARS.getLDAPLinuxAccessGroupMembers(baseDN_MARS_AccessGroup, searchScope_MARS_AccessGroup,
                                                                                                    searchFilter_App_MARS_AccessGroup,
                                                                                                    retrieveAttributes_MARS_AccessGroup, derefAliases_MARS_AccessGroup)
  marsLinuxAccessGroupMembers_l.append(stg_auditOSHSecurityComplianceMARS.getLDAPLinuxAccessGroupMembers(baseDN_MARS_AccessGroup, searchScope_MARS_AccessGroup,
                                                                                                         searchFilter_DB_MARS_AccessGroup,
                                                                                                         retrieveAttributes_MARS_AccessGroup, derefAliases_MARS_AccessGroup))
  # pp.pprint(marsLinuxAccessGroupMembers_l)
  df_secures = pd.DataFrame({'MARS': marsLinuxAccessGroupMembers_l})
  df_secures.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Access", index=False)

  stg_auditOSHSecurityComplianceMARS.wb2.save()

  ## MARS - Linux groups:
  baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  searchScope_MARS_LinuxGroups = SUBTREE
  gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  derefAliases_MARS_LinuxGroups = DEREF_ALWAYS

  gidMembers_d = stg_auditOSHSecurityComplianceMARS.getLinuxGroupMembers(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
                                                                         retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)

  for grpID in gidMembers_d.keys():

    print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceMARS.DateTime, grpID))

    countGroupMembers = 0
    for member in gidMembers_d[grpID]:

      print(member)
      countGroupMembers += 1

    print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))

  df_secures = pd.DataFrame({'MARS': marsLinuxAccessGroupMembers_l})
  df_secures.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Groups", index=False)

  stg_auditOSHSecurityComplianceMARS.wb2.save()



  #############################################################################################################
  #
  #
  #############################################################################################################
  ## Polling - Access Groups:
  baseDN_MARS_AccessGroup = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  searchScope_MARS_AccessGroup = SUBTREE
  searchFilter_App_MARS_AccessGroup = '(cn=Linux-MARS-App-Access*)'
  searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  retrieveAttributes_MARS_AccessGroup = ['cn', 'member']
  derefAliases_MARS_AccessGroup = DEREF_ALWAYS

  curPolling = '00001'
  nextPolling = str( int(curPolling) + 1 )  # To do: should detect the number from existing filename_sample and increment by 1
  pathPolling = r"C:/staging/python/pilot/testPolling Account Review " + nextPolling + r".xlsx"


  stg_auditOSHSecurityCompliancePolling = auditSecComp_ldap(ldapServer, userDN, password, pathPolling)

  pollingLinuxAccessGroupMembers_l = stg_auditOSHSecurityCompliancePolling.getLDAPLinuxAccessGroupMembers(baseDN_MARS_AccessGroup, searchScope_MARS_AccessGroup,
                                                                                                          searchFilter_App_MARS_AccessGroup,
                                                                                                          retrieveAttributes_MARS_AccessGroup, derefAliases_MARS_AccessGroup)
  pollingLinuxAccessGroupMembers_l.append(stg_auditOSHSecurityCompliancePolling.getLDAPLinuxAccessGroupMembers(baseDN_MARS_AccessGroup, searchScope_MARS_AccessGroup,
                                                                                                               searchFilter_DB_MARS_AccessGroup,
                                                                                                               retrieveAttributes_MARS_AccessGroup, derefAliases_MARS_AccessGroup))
  # pp.pprint(marsLinuxAccessGroupMembers_l)
  df_secures = pd.DataFrame({'Polling': pollingLinuxAccessGroupMembers_l})
  df_secures.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Access", index=False)

  stg_auditOSHSecurityCompliancePolling.wb2.save()

  ## MARS - Linux groups:
  baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  searchScope_MARS_LinuxGroups = SUBTREE
  gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  derefAliases_MARS_LinuxGroups = DEREF_ALWAYS

  gidMembers_d = stg_auditOSHSecurityCompliancePolling.getLinuxGroupMembers(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
                                                                            retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)

  PollingMembers_l = []
  for grpID in gidMembers_d.keys():

    print('{} - Linux Group ID [{}]:'.format(Enums.strDateTime, grpID))

    countGroupMembers = 0
    for member in gidMembers_d[grpID]:

      # print(member)
      PollingMembers_l.append(member)
      countGroupMembers += 1

    print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))


  df_secures = pd.DataFrame({'Polling': PollingMembers_l})
  df_secures.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Groups", index=False)

  stg_auditOSHSecurityCompliancePolling.wb2.save()





