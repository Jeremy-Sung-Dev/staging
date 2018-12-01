#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" auditSoxCompliance_Ldap.py

Description:
Attributes:
  __version__ = "0.0.1"
  __project__ = Staging
  __author__ = Jeremy Sung
  __date__ = 6/20/2018 3:36 PM
  __Email__ = jsung8@gmail.com

Todo:
"""

import os, re
from ldap3 import Server, Connection, SUBTREE, DEREF_ALWAYS
import Enums
import pandas as pd
from openpyxl import Workbook, load_workbook
import bz2
import paramiko
import pprint as pp
import argparse




class pilot_pySSH:

  def __init__(self, hostname="", port=22, userid="", passwd="", privateKey=""):
    self.hostname = hostname
    self.port = port if port else 22
    self.userid = userid
    self.passwd = passwd if passwd else ""
    self.privateKey = privateKey

    self.ssh = paramiko.SSHClient()
    self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    if privateKey:
      self.ssh.connect(self.hostname, self.port, username=self.userid, key_filename=self.privateKey)
    else:
      self.ssh.connect(self.hostname, self.port, username=self.userid, password=self.passwd)


  def getSudoers(self, cmd):

    # ## Ref.  https://stackoverflow.com/questions/10745138/python-paramiko-ssh
    # server = '10.1.19.21'  # wms-app91.orchard.osh
    # port = 22
    # userid = 'root'

    # privateKey_root = "C:/admin/id_rsa_root"
    ## ssh.connect(server, port, user, passwd)  # Worked!!
    ## ssh.connect(server, port, username=user, password=passwd, key_filename="C:/admin/id_rsa_root") # Private Key!!

    # auditSC_LocalAdminAccess = auditSoxCompliance_ServerLocal(server=server, port=port, userid=userid, privateKey=privateKey_root)

    ## If "sudo" is required the script need to promput user for password though:
    # cmd = "echo {} | sudo -S {}".format(",.SOLpx.01", "cat /etc/sudoers")

    stdin, stdout, stderr = self.ssh.exec_command(cmd)

    result_l = []

    for line in stdout.readlines():
      # if '^#' in line or '^$' in line:
      #   continue
      result_l.append(line)

    # pp.pprint(result_l)

    self.ssh.close()
    return result_l





class stg_auditOSHSecurityCompliance:

  def __init__(self, ldapServer, userDN, password, reportPath):
    Domain = '.orchard.osh'
    if '.' not in ldapServer:
      self.ldapServer = ldapServer + Domain

    self.ldapServer = "dc01.orchard.osh" if not ldapServer else ldapServer
    self.UserDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh" if not userDN else userDN
    self.Password = "Ld@p$3cAuth!" if not password else password
    ## self.Port = 636

    ## first open a connection to the server:
    server = Server(self.ldapServer)
    self.conn = Connection(server, user=self.UserDN, password=self.Password)

    self.lastMonth = Enums.LastMonth  # May
    self.DateTime = Enums.DateTime

    if not os.path.isfile(reportPath):
      self.wb2 = Workbook()
      # self.wb2.remove_sheet(self.wb2.active)
      self.wb2.save(reportPath)

    self.report = reportPath      # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

    self.wb2 = pd.ExcelWriter(self.report, engine='openpyxl')
    # self.wb2.remove_sheet(self.wb2.active)
    self.book = load_workbook(self.report)
    self.wb2.book = self.book

    # Remove the default sheet
    default_sheet = self.wb2.book["Sheet"]
    self.wb2.book.remove(default_sheet)
    # default_sheet = self.wb2.book.get_sheet_by_name("Sheet") # deprecated function get_sheet_by_name (Use wb[sheetname])
    # self.wb2.book.remove_sheet(default_sheet) # deprecated function remove_sheet (Use wb.remove(worksheet) or del wb[sheetname])


  def getLDAPLinuxAccessGroupMembers(self, baseDN, searchScope, searchFilter, retrieveAttributes, derefAliases):

    self.conn.open()
    self.conn.bind()

    wmsLinuxAccessGroupMembers_l = []

    derefAliases = DEREF_ALWAYS if not derefAliases else derefAliases

    ## Review WMS Access Groups - Print group members for server access.
    self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter, attributes=retrieveAttributes,
                dereference_aliases=derefAliases)

    ## WMS LDAP access:
    responses_d = self.conn.response[0]
    attributes_d = responses_d['attributes']

    for member in attributes_d['member']:
      wmsLinuxAccessGroupMembers_l.append(member)

    self.conn.unbind()
    return wmsLinuxAccessGroupMembers_l


  def getLinuxGroupMembers(self, baseDN, searchScope, retrieveAttributes, gidNumbers):

    self.conn.open()
    self.conn.bind()

    gidMembers_d = {}
    for gidNum in gidNumbers:
      searchFilter = '(gidNumber=' + gidNum + ')'

      self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter,
                  attributes=retrieveAttributes)

      gidMembers_l = []
      for member_d in self.conn.response:
        for key, value in member_d.items():
          if key == 'dn':
            gidMembers_l.append(member_d[key])
          else:
            continue

      gidMembers_d[gidNum] = gidMembers_l

    self.conn.unbind()
    return gidMembers_d


  ## WMS - Linux Audit Hosts - Verify local Server Accounts and Sudo file:
  # auditHosts = ['wms-wmosapp01', 'wms-wmosdb01', 'wms-sciapp01', 'wms-scidb01', 'wms-ifeeapp01']


  def auditAdminAccountsSolo_Helper(self, filename, excludedPhrases):

    entries_l = []

    with open(filename, 'r') as file:

      entries_l.append("Date:\t{}\n".format(Enums.DateTime))

      for entry in file:
        if any(term in entry for term in excludedPhrases):
          continue
        entries_l.append(entry.strip())

    return entries_l


  def auditAdminAccountsSolo(self, filename):
    """ Expecting: /etc/passwd, /etc/sudoers, or /etc/shadow """

    # server = "host_name"
    # CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
    # CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"
    CMD_passwd_shadow = "\nCommand:\t{}".format("date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1")
    CMD_sudoers = "\nCommand:\t{}".format("date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'")

    Excluded_passwd = ('nologin', 'false')
    Excluded_shadow = (':!', ':*')
    Excluded_sudoers = ('^#', '^$')

    resPassword_l = []
    resShadow_l = []
    resSudoers_l = []

    if filename.startswith("passwd"):
      # resPassword_l.append(server)
      resPassword_l.append(CMD_passwd_shadow)
      resPassword_l = self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd)
      return resPassword_l

    # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
    if filename.startswith("shadow"):
      # resShadow_l.append(server)
      resShadow_l.append(CMD_passwd_shadow)
      resShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))
      return resShadow_l

    if filename.startswith("sudoers"):
      # resSudoers_l.append(server)
      resSudoers_l.append(CMD_sudoers)
      resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
      return resSudoers_l

    else:
      print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename))


  def auditAdminAccounts(self, *args):
    """ args represent a tuple or a list of files, such as /etc/passwd, shadow or sudoers """

    if not args:
      print("No file presents. Try again.\n")

    auditResult_l = []
    for filename in args:
      # auditResult_l.extend(self.scanLocalAdminAccounts_Server_Solo(filename_sample))

      # server = "host_name"
      CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
      CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"

      Excluded_passwd = ('nologin', 'false')
      Excluded_shadow = (':!', ':*')
      Excluded_sudoers = ('^#', '^$')

      resPassword_l = []
      resPasswordShadow_l = []
      resSudoers_l = []

      if filename.startswith("passwd"):

        # resPassword_l.append(server)
        resPassword_l.append(CMD_passwd_shadow)
        resPassword_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd))

        # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
        if resPassword_l and filename.startswith("shadow"):
          # resPasswordShadow_l.append(CMD_passwd_shadow)
          resPasswordShadow_l.extend(resPassword_l)
          resPasswordShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))

          # return resPasswordShadow_l
          auditResult_l.extend(resPasswordShadow_l)


        else:
          # return resPassword_l
          auditResult_l.extend(resPassword_l)


      if filename.startswith("sudoers"):
        # resSudoers_l.append(server)
        resSudoers_l.append(CMD_sudoers)
        resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
        # return resSudoers_l
        auditResult_l.extend(resSudoers_l)

      # else:
      #   print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename_sample))

    return auditResult_l


  def auditHostAccess_ServerLocalSecureMessages(self):
    """ /etc/secure and /etc/secure.* """

    print(self.DateTime)

    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')

    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            # words =  line.split()
            # reduced_line = " ".join([w for w in words if w not in Excluded])

            # if not any(term in line for term in Excluded):
            if any(term in line for term in Excluded):
              continue

            groups = compiled.search(line)
            # print(groups.group(1).strip())  # May 15 10:16:19
            # print(groups.group(2).strip())  # hostnames
            # print(groups.group(3).strip())  # Messages

            # timestamp, host, message = groups.group(1).strip(), groups.group(2).strip(), groups.group(3)
            # print("{}, {}, {}".format(timestamp, host, message))
            # print("{} -- {} -- {}".format(timestamps, host, messages))

            timestamps_l.append(groups.group(1).strip())
            hostnames_l.append(groups.group(2).strip())
            messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Host Access", index=False)

    self.wb2.save()


  def audit_OracleAccountAccess_DFIODB01_Secure(self):
    """ /etc/secure.* """

    print(self.DateTime)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    # Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')
    Included = ('sudo', 'USER=oracle')
    path = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            groups = compiled.search(line)

            if any(term in groups.group(3) for term in Included):
              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Oracle Account Access", index=False)


    self.wb2.save()



  def auditInformixAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.DateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "informix"]
    # Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()



            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))

    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Informix Access", index=False)


    self.wb2.save()




  def auditHostRootAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.DateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "sshd", "root"]
    Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []


    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            # print(line) # b'Jun 17 17:30:01 polling61 /usr/sbin/cron[25801]: (root) CMD ([ -x /usr/lib64/sa/sa1 ] && exec /usr/lib64/sa/sa1 -S ALL 1 1)\n'

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()

            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    # df_pollingLinuxAccess.to_excel(writer, sheet_name="Root Access", index=False)
    df_secures.to_excel(self.wb2, sheet_name="Root Access", index=False)

    self.wb2.save()



if __name__ == "__main__":


  ###############################################################################################
  ## Argparse Initiatives:
  parser = argparse.ArgumentParser(prog='dev_oshr.py', prefix_chars='-+', description='util to execute Linux commands',
                                   add_help=True, allow_abbrev=True)
  parser.add_argument("-C", "--command", required=True, help="Linux command to execute")
  # parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # parser.add_argument("-C", "--command", action="store_true", help="Linux command to execute")
  ## Positional:
  # parser.add_argument('bar', help='positional bar')
  parser.add_argument("--reportPath", default="C:/staging/python/pilot", help="Show content in reportPath")
  # pp.pprint(parser)

  args = parser.parse_args()
  # pp.pprint(args)

  ## Have to execute with "python" in front!!
  ## C:\staging\python\pilot>python dev_oshr.py --command date
  ## Namespace(command='date', reportPath='C:/staging/python/pilot')

  ## Doesn't work:
  ## C:\staging\python\pilot> dev_oshr.py --command date  # unable to recognize "--command date"


  # print(args.command)
  # print("Command: {}".format(args.command))
  print("Command: {},\tReport: {}".format(args.command, args.report))
  # pp.pprint(args.command)
  # result_l = auditSC_LocalAdminAccess.exeC(args.co)

  # pp.pprint(result_l)

  ###############################################################################################

  # sol = auditSoxCompliance_Ldap()

  # ldapServer = "dc01.orchard.osh"
  # userDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh"
  # password = "Ld@p$3cAuth!"
  #
  # #############################################################################################################
  #
  #
  # # pathNewWMSReport = r"C:/staging/python/pilot/testWMS Account Review " + nextWMSNum + r".xlsx"
  # pathNewWMSReport = r"C:/staging/python/pilot/testWMS Account Review.xlsx"
  # stg_auditOSHSecurityComplianceWMS = auditSoxCompliance_Ldap(ldapServer, userDN, password, pathNewWMSReport)
  #
  # ## WMS - Access Groups:
  # baseDN_WMS_LdapAccessGroups = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  # searchScope_WMS_LdapAccessGroups = SUBTREE
  # searchFilter_WMS_LdapAccessGroups = '(cn=Linux-WMS-Access*)'
  # retrieveAttributes_WMS_LdapAccessGroups = ['cn', 'member']
  # derefAliases_WMS_LdapAccessGroups = DEREF_ALWAYS
  #
  # wmsLinuxAccess_l = stg_auditOSHSecurityComplianceWMS.auditLinuxAdminLdapAccess(baseDN_WMS_LdapAccessGroups, searchScope_WMS_LdapAccessGroups,
  #                                                                                                 searchFilter_WMS_LdapAccessGroups,
  #                                                                                                 retrieveAttributes_WMS_LdapAccessGroups, derefAliases_WMS_LdapAccessGroups)
  # # pp.pprint(wmsLinuxAccess_l)
  #
  # df_pollingLinuxAccess = pd.DataFrame({'WMS': wmsLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceWMS.wb2, sheet_name="WMS LDAP Access", index=False)
  #
  # stg_auditOSHSecurityComplianceWMS.wb2.save()
  #
  # ## WMS - Linux groups:
  # baseDN_WMS_LinuxGroups = "dc=orchard,dc=osh"
  # searchScope_WMS_LinuxGroups = SUBTREE
  # gidNumbers_WMS_LinuxGroups = ['10000', '10001', '10004', '10005']
  # retrieveAttributes_WMS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  # derefAliases_WMS_LinuxGroups = DEREF_ALWAYS
  #
  # gidMembers_d = stg_auditOSHSecurityComplianceWMS.listLdapAdminGroupMembership(baseDN_WMS_LinuxGroups, searchScope_WMS_LinuxGroups,
  #                                                                       retrieveAttributes_WMS_LinuxGroups, gidNumbers_WMS_LinuxGroups)
  # wmsLdapAdminGroups_l = []
  # for grpID in gidMembers_d.keys():
  #
  #   # print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceWMS.DateTime, grpID))
  #   wmsLdapAdminGroups_l.append('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceWMS.DateTime, grpID))
  #
  #   countGroupMembers = 0
  #   for member in gidMembers_d[grpID]:
  #     # print(member)
  #     wmsLdapAdminGroups_l.append(member)
  #     countGroupMembers += 1
  #
  #   # print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #   wmsLdapAdminGroups_l.append('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #
  # df_pollingLinuxAccess = pd.DataFrame({'WMS': wmsLdapAdminGroups_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceWMS.wb2, sheet_name="WMS LDAP Groups", index=False)
  #
  # stg_auditOSHSecurityComplianceWMS.wb2.save()
  #
  # stg_auditOSHSecurityComplianceWMS.scanLocalAdminAccounts_Local_Files()
  # stg_auditOSHSecurityComplianceWMS.scanHostAccess_Local_VarLogSecure()
  #
  #
  # #############################################################################################################
  #
  #
  # #############################################################################################################
  # ## MARS - Access Groups:
  # baseDN_MARS_LdapAccessGroups = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  # searchScope_MARS_LdapAccessGroup = SUBTREE
  # searchFilter_App_MARS_LdapAccessGroup = '(cn=Linux-MARS-App-Access*)'
  # searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  # retrieveAttributes_MARS_LdapAccessGroup = ['cn', 'member']
  # derefAliases_MARS_LdapAccessGroup = DEREF_ALWAYS
  #
  # pathMARS = r"C:/staging/python/pilot/testMARS Account Review.xlsx"
  # stg_auditOSHSecurityComplianceMARS = auditSoxCompliance_Ldap(ldapServer, userDN, password, pathMARS)
  #
  # marsLinuxAccess_l = stg_auditOSHSecurityComplianceMARS.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                                   searchFilter_App_MARS_LdapAccessGroup,
  #                                                                                                   retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup)
  # marsLinuxAccess_l.append(stg_auditOSHSecurityComplianceMARS.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                                        searchFilter_DB_MARS_AccessGroup,
  #                                                                                                        retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup))
  # # pp.pprint(marsLinuxAccess_l)
  # df_pollingLinuxAccess = pd.DataFrame({'MARS': marsLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Access", index=False)
  #
  # stg_auditOSHSecurityComplianceMARS.wb2.save()
  #
  # ## MARS - Linux groups:
  # baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  # searchScope_MARS_LinuxGroups = SUBTREE
  # gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  # retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  # derefAliases_MARS_LinuxGroups = DEREF_ALWAYS
  #
  # gidMembers_d = stg_auditOSHSecurityComplianceMARS.listLdapAdminGroupMembership(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
  #                                                                        retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)
  #
  # for grpID in gidMembers_d.keys():
  #
  #   print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceMARS.DateTime, grpID))
  #
  #   countGroupMembers = 0
  #   for member in gidMembers_d[grpID]:
  #
  #     print(member)
  #     countGroupMembers += 1
  #
  #   print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #
  # df_pollingLinuxAccess = pd.DataFrame({'MARS': marsLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Groups", index=False)
  #
  # stg_auditOSHSecurityComplianceMARS.wb2.save()
  #
  # stg_auditOSHSecurityComplianceMARS.scanLocalAdminAccounts_Local_Files()
  # stg_auditOSHSecurityComplianceMARS.scanHostAccess_Local_VarLogSecure()
  # stg_auditOSHSecurityComplianceMARS.scanHostAccess_Local_VarLogSecure_Oracle()
  #
  # #############################################################################################################
  # #
  # #
  # #############################################################################################################
  # ## Polling - Access Groups:
  # baseDN_MARS_LdapAccessGroups = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  # searchScope_MARS_LdapAccessGroup = SUBTREE
  # searchFilter_App_MARS_LdapAccessGroup = '(cn=Linux-MARS-App-Access*)'
  # searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  # retrieveAttributes_MARS_LdapAccessGroup = ['cn', 'member']
  # derefAliases_MARS_LdapAccessGroup = DEREF_ALWAYS
  #
  # pathPolling = r"C:/staging/python/pilot/testPolling Account Review.xlsx"
  # stg_auditOSHSecurityCompliancePolling = auditSoxCompliance_Ldap(ldapServer, userDN, password, pathPolling)
  #
  # pollingLinuxAccess_l = stg_auditOSHSecurityCompliancePolling.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                                         searchFilter_App_MARS_LdapAccessGroup,
  #                                                                                                         retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup)
  # pollingLinuxAccess_l.append(stg_auditOSHSecurityCompliancePolling.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                                              searchFilter_DB_MARS_AccessGroup,
  #                                                                                                              retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup))
  # # pp.pprint(marsLinuxAccess_l)
  # df_pollingLinuxAccess = pd.DataFrame({'Polling': pollingLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Access", index=False)
  #
  # stg_auditOSHSecurityCompliancePolling.wb2.save()
  #
  # ## MARS - Linux groups:
  # baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  # searchScope_MARS_LinuxGroups = SUBTREE
  # gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  # retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  # derefAliases_MARS_LinuxGroups = DEREF_ALWAYS
  #
  # gidMembers_d = stg_auditOSHSecurityCompliancePolling.listLdapAdminGroupMembership(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
  #                                                                           retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)
  #
  # pollingLdapAdminGroups_l = []
  # for grpID in gidMembers_d.keys():
  #
  #   print('{} - Linux Group ID [{}]:'.format(Enums.DateTime, grpID))
  #
  #   countGroupMembers = 0
  #   for member in gidMembers_d[grpID]:
  #
  #     # print(member)
  #     pollingLdapAdminGroups_l.append(member)
  #     countGroupMembers += 1
  #
  #   print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #
  #
  # df_pollingLinuxAccess = pd.DataFrame({'Polling': pollingLdapAdminGroups_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Groups", index=False)
  #
  # stg_auditOSHSecurityCompliancePolling.wb2.save()
  #
  # stg_auditOSHSecurityCompliancePolling.scanLocalAdminAccounts_Local_Files()
  # stg_auditOSHSecurityCompliancePolling.scanHostAccess_Local_VarLogMessages_BZ2_Root_Polling()
  # stg_auditOSHSecurityCompliancePolling.scanHostAccess_Local_VarLogMessages_BZ2_Informix_Polling()

  #############################################################################################################


  # ldapServer = "dc01.orchard.osh"
  # userDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh"
  # passwd = "Ld@p$3cAuth!"
  #
  # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"
  #
  # stg_auditOSHSecurityCompliance_ServerLocalAccess = auditSoxCompliance_Ldap(ldapServer, userDN, passwd, reportPath)
  # # stg_auditOSHSecurityCompliance_ServerLocalAccess = auditSoxCompliance_Ldap(reportPath)
  #
  # # ## /etc/passwd, shadow, sudoers:
  # # filename1 = "passwd"
  # # filename2 = "shadow"
  # # filename3 = "sudoers"
  # # result_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.scanLocalAdminAccounts_Server_Solo(filename1)
  # # filenames = [filename1, filename2, filename3]
  # # auditResults_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.scanLocalAdminAccounts_Local_Files(*filenames)
  # # for entry in auditResults_l:
  # #   print(entry)
  #
  # # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogSecure()
  #
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogSecure()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogSecure_Oracle()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogMessages_BZ2_Informix_Polling()
  # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogMessages_BZ2_Root_Polling()

  ##############################################################################################
  ## Worked!
  ## Can run any Linux commands as root@admin01
  ## Get /etc/sudoers as root

  # Ref.  https://stackoverflow.com/questions/10745138/python-paramiko-ssh
  # server = '10.1.19.21'  # wms-app91.orchard.osh
  hostname = '10.1.2.210'  # wms-wmosapp01.orchard.osh
  port = 22
  userid = 'root'

  privateKey_root = "C:/admin/id_rsa_root"
  # ssh.connect(server, port, user, passwd)  # Worked!!
  # ssh.connect(server, port, username=user, password=passwd, key_filename="C:/admin/id_rsa_root") # Private Key!!

  pilotSSH = pilot_pySSH(hostname=hostname, port=port, userid=userid, privateKey=privateKey_root)

  # cmd_sudoers = 'cat /etc/sudoers'
  # cmd_passwd = 'cat /etc/passwd'
  ## If "sudo" is required the script need to promput user for password though:
  # cmd = "echo {} | sudo -S {}".format(",.SOLpx.01", "cat /etc/sudoers")
  # cmd = 'cat /var/log/secure | head -30'
  cmd = "server; cat /etc/*release; date"

  # stdin, stdout, stderr = auditSC_LocalAdminAccess.exeC(cmd)

  # result_l = auditSC_LocalAdminAccess.exeC(cmd_passwd)
  result_l = pilotSSH.getSudoers(cmd)

  pp.pprint(result_l)

