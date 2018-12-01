#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" dev_osh_audit_ServerLocal_Access.py

Challenges:
Solutions:
Description:
Attributes:
  __version__ = "0.0.0"
  __project__ = pilot
  __author__ = Jeremy Sung
  __date__ = 6/12/2018 4:19 PM
  __Email__ = jsung8@gmail.com
Todo:
"""

import os, re
import pandas as pd
import Enums
from openpyxl import Workbook, load_workbook
import bz2


class osh_audit_ServerLocal_Access:

  def __init__(self, path):

    self.lastMonth = Enums.LastMonth  # May
    self.DateTime = Enums.DateTime

    if not os.path.isfile(path):
      self.wb2 = Workbook()
      self.wb2.save(path)

    self.path = path      # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

    self.wb2 = pd.ExcelWriter(self.path, engine='openpyxl')
    self.book = load_workbook(self.path)
    self.wb2.book = self.book


  ## WMS - Linux Audit Hosts - Verify local Server Accounts and Sudo file:
  # auditHosts = ['wms-wmosapp01', 'wms-wmosdb01', 'wms-sciapp01', 'wms-scidb01', 'wms-ifeeapp01']


  def auditAdminAccountsSolo_Helper(self, filename, excludedPhrases):

    entries_l = []

    with open(filename, 'r') as file:

      entries_l.append("Date:\t{}\n".format(Enums.DateTime))

      for entry in file:
        if any(term in entry for term in excludedPhrases):
          continue
        entries_l.append(entry.strip())

    return entries_l


  def auditAdminAccountsSolo(self, filename):
    """ Expecting: /etc/passwd, /etc/sudoers, or /etc/shadow """

    # server = "host_name"
    # CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
    # CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"
    CMD_passwd_shadow = "\nCommand:\t{}".format("date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1")
    CMD_sudoers = "\nCommand:\t{}".format("date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'")

    Excluded_passwd = ('nologin', 'false')
    Excluded_shadow = (':!', ':*')
    Excluded_sudoers = ('^#', '^$')

    resPassword_l = []
    resShadow_l = []
    resSudoers_l = []

    if filename.startswith("passwd"):
      # resPassword_l.append(server)
      resPassword_l.append(CMD_passwd_shadow)
      resPassword_l = self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd)
      return resPassword_l

    # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
    if filename.startswith("shadow"):
      # resShadow_l.append(server)
      resShadow_l.append(CMD_passwd_shadow)
      resShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))
      return resShadow_l

    if filename.startswith("sudoers"):
      # resSudoers_l.append(server)
      resSudoers_l.append(CMD_sudoers)
      resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
      return resSudoers_l

    else:
      print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename))


  def auditAdminAccounts(self, *args):
    """ args represent a tuple or a list of files, such as /etc/passwd, shadow or sudoers """

    if not args:
      print("No file presents. Try again.\n")

    auditResult_l = []
    for filename in args:
      # auditResult_l.extend(self.scanLocalAdminAccounts_Server_Solo(filename_sample))

      # server = "host_name"
      CMD_passwd_shadow = "\nCommand: date && cat /etc/passwd | grep -v 'nologin' | grep -v 'false' && cat /etc/shadow | grep -v ':!' | grep -v ':\*' | cut -d ':' -f 1"
      CMD_sudoers = "\nCommand: date && cat /etc/sudoers | grep -v '^#' | grep -v '^$'"

      Excluded_passwd = ('nologin', 'false')
      Excluded_shadow = (':!', ':*')
      Excluded_sudoers = ('^#', '^$')

      resPassword_l = []
      resPasswordShadow_l = []
      resSudoers_l = []

      if filename.startswith("passwd"):

        # resPassword_l.append(server)
        resPassword_l.append(CMD_passwd_shadow)
        resPassword_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_passwd))

        # if len(resPassword_l) > 0 and filename_sample.startswith("shadow"):
        if resPassword_l and filename.startswith("shadow"):
          # resPasswordShadow_l.append(CMD_passwd_shadow)
          resPasswordShadow_l.extend(resPassword_l)
          resPasswordShadow_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_shadow))

          # return resPasswordShadow_l
          auditResult_l.extend(resPasswordShadow_l)


        else:
          # return resPassword_l
          auditResult_l.extend(resPassword_l)


      if filename.startswith("sudoers"):
        # resSudoers_l.append(server)
        resSudoers_l.append(CMD_sudoers)
        resSudoers_l.extend(self.auditAdminAccountsSolo_Helper(filename, Excluded_sudoers))
        # return resSudoers_l
        auditResult_l.extend(resSudoers_l)

      # else:
      #   print("File {} is not recognized. Please enter: /etc/passwd, shadow or sudoers.\n".format(filename_sample))

    return auditResult_l


  def auditHostAccess_ServerLocalSecureMessages(self):
    """ /etc/secure and /etc/secure.* """

    print(self.DateTime)

    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')

    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            # words =  line.split()
            # reduced_line = " ".join([w for w in words if w not in Excluded])

            # if not any(term in line for term in Excluded):
            if any(term in line for term in Excluded):
              continue

            groups = compiled.search(line)
            # print(groups.group(1).strip())  # May 15 10:16:19
            # print(groups.group(2).strip())  # hostnames
            # print(groups.group(3).strip())  # Messages

            # timestamp, host, message = groups.group(1).strip(), groups.group(2).strip(), groups.group(3)
            # print("{}, {}, {}".format(timestamp, host, message))
            # print("{} -- {} -- {}".format(timestamps, host, messages))

            timestamps_l.append(groups.group(1).strip())
            hostnames_l.append(groups.group(2).strip())
            messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Host Access", index=False)

    self.wb2.save()


  def audit_OracleAccountAccess_DFIODB01_Secure(self):
    """ /etc/secure.* """

    print(self.DateTime)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    # Excluded = ('rfuser', 'disconnect', 'pam_unix', 'subsystem request for sftp')
    Included = ('sudo', 'USER=oracle')
    path = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    for filename in os.listdir("."):

      if filename.startswith("secure"):

        with open(filename, 'r') as secure:

          for line in secure:

            if not line.startswith(self.lastMonth):
              continue

            groups = compiled.search(line)

            if any(term in groups.group(3) for term in Included):
              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Oracle Account Access", index=False)


    self.wb2.save()



  def auditInformixAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.DateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "informix"]
    # Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []

    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()



            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))

    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    df_secures.to_excel(self.wb2, sheet_name="Informix Access", index=False)


    self.wb2.save()




  def auditHostRootAccess_Polling_ServerLocalMessages_BZ2(self):
    """ /etc/messages-*  bz2  """

    # print(Enums.DateTime)
    print(self.lastMonth)


    pattern = re.compile("(\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2})\s+(\S+)\s+(.*$)")
    compiled = re.compile(pattern)

    Included = ["sudo", "sshd", "root"]
    Excluded = ("nagios")


    timestamps_l = []
    hostnames_l = []
    messages_l = []


    # Ref. https://pymotw.com/3/bz2/

    for filename in os.listdir("."):

      if filename.startswith("messages"):


        with bz2.open(filename, 'rt', encoding = "utf-8") as input:


          for line in input:

            # print(line) # b'Jun 17 17:30:01 polling61 /usr/sbin/cron[25801]: (root) CMD ([ -x /usr/lib64/sa/sa1 ] && exec /usr/lib64/sa/sa1 -S ALL 1 1)\n'

            if not line.startswith(self.lastMonth):
              continue

            words =  line.split()

            groups = compiled.search(line)

            if any(term in line for term in Included):

              timestamps_l.append(groups.group(1).strip())
              hostnames_l.append(groups.group(2).strip())
              messages_l.append(groups.group(3))


    df_secures = pd.DataFrame({'TimeStamp': timestamps_l, 'Hosts': hostnames_l, 'Messages': messages_l})
    df_secures = df_secures[["TimeStamp", "Hosts", "Messages"]]
    # df_pollingLinuxAccess.to_excel(writer, sheet_name="Root Access", index=False)
    df_secures.to_excel(self.wb2, sheet_name="Root Access", index=False)

    self.wb2.save()







if __name__ == "__main__":


  path = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

  oshServerLocalAccess = osh_audit_ServerLocal_Access(path)

  # ## /etc/passwd, shadow, sudoers:
  # filename1 = "passwd"
  # filename2 = "shadow"
  # filename3 = "sudoers"
  # result_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.scanLocalAdminAccounts_Server_Solo(filename1)
  # filenames = [filename1, filename2, filename3]
  # auditResults_l = stg_auditOSHSecurityCompliance_ServerLocalAccess.scanLocalAdminAccounts_Local_Files(*filenames)
  # for entry in auditResults_l:
  #   print(entry)

  # stg_auditOSHSecurityCompliance_ServerLocalAccess.scanHostAccess_Local_VarLogSecure()

  oshServerLocalAccess.auditHostAccess_ServerLocalSecureMessages()
  oshServerLocalAccess.audit_OracleAccountAccess_DFIODB01_Secure()
  oshServerLocalAccess.auditInformixAccess_Polling_ServerLocalMessages_BZ2()
  oshServerLocalAccess.auditHostRootAccess_Polling_ServerLocalMessages_BZ2()
  
  