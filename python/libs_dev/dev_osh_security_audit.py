#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" dev_osh_security_audit.py

Description:
Attributes:
  __version__ = "0.0.0"
  __project__ = pilot
  __author__ = Jeremy Sung
  __date__ = 6/12/2018 8:47 AM
  __Email__ = jsung8@gmail.com
Todo:
"""

import os
import pprint as pp
from ldap3 import Server, Connection, SUBTREE, DEREF_ALWAYS
import Enums
import pandas as pd
from openpyxl import Workbook, load_workbook


class osh_security_audit:

  def __init__(self, ldapServer, userDN, password, path):
    Domain = '.orchard.osh'
    if '.' not in ldapServer:
      self.ldapServer = ldapServer + Domain

    self.ldapServer = "dc01.orchard.osh" if not ldapServer else ldapServer
    self.UserDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh" if not userDN else userDN
    self.Password = "Ld@p$3cAuth!" if not password else password
    ## self.Port = 636

    ## first open a connection to the server:
    server = Server(self.ldapServer)
    self.conn = Connection(server, user=self.UserDN, password=self.Password)

    self.lastMonth = Enums.LastMonth  # May
    self.DateTime = Enums.DateTime

    if not os.path.isfile(path):
      self.wb2 = Workbook()
      self.wb2.save(path)

    self.path = path      # reportPath = r"C:/staging/python/pilot/test_fromBZ2_all.xlsx"

    self.wb2 = pd.ExcelWriter(self.path, engine='openpyxl')
    self.book = load_workbook(self.path)
    self.wb2.book = self.book


  def getLDAPLinuxAccessGroupMembers(self, baseDN, searchScope, searchFilter, retrieveAttributes, derefAliases):

    self.conn.open()
    self.conn.bind()

    wmsLinuxAccessGroupMembers_l = []

    derefAliases = DEREF_ALWAYS if not derefAliases else derefAliases

    ## Review WMS Access Groups - Print group members for server access.
    self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter, attributes=retrieveAttributes,
                dereference_aliases=derefAliases)

    ## WMS LDAP access:
    responses_d = self.conn.response[0]
    attributes_d = responses_d['attributes']

    for member in attributes_d['member']:
      wmsLinuxAccessGroupMembers_l.append(member)

    self.conn.unbind()
    return wmsLinuxAccessGroupMembers_l


  def getLinuxGroupMembers(self, baseDN, searchScope, retrieveAttributes, gidNumbers):

    self.conn.open()
    self.conn.bind()

    gidMembers_d = {}
    for gidNum in gidNumbers:
      searchFilter = '(gidNumber=' + gidNum + ')'

      self.conn.search(search_base=baseDN, search_scope=searchScope, search_filter=searchFilter,
                  attributes=retrieveAttributes)

      gidMembers_l = []
      for member_d in self.conn.response:
        for key, value in member_d.items():
          if key == 'dn':
            gidMembers_l.append(member_d[key])
          else:
            continue

      gidMembers_d[gidNum] = gidMembers_l

    self.conn.unbind()
    return gidMembers_d







if __name__ == "__main__":

  ldapServer = "dc01.orchard.osh"
  userDN = "CN=Lookitup4,OU=Users,OU=Administrative,DC=Orchard,DC=osh"
  password = "Ld@p$3cAuth!"

  #############################################################################################################
  pathWMS = r"C:/staging/python/pilot/testWMSAccountReview.xlsx"
  oshSecurityAuditWMS = osh_security_audit(ldapServer, userDN, password, pathWMS)

  ## WMS - Access Groups:
  baseDN_WMS_AccessGroup = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  searchScope_WMS_AccessGroup = SUBTREE
  searchFilter_WMS_AccessGroup = '(cn=Linux-WMS-Access*)'
  retrieveAttributes_WMS_AccessGroup = ['cn', 'member']
  derefAliases_WMS_AccessGroup = DEREF_ALWAYS

  wmsLinuxAccessGroupMembers_l = oshSecurityAuditWMS.getLDAPLinuxAccessGroupMembers(baseDN_WMS_AccessGroup, searchScope_WMS_AccessGroup,
                                                                                 searchFilter_WMS_AccessGroup,
                                                                                 retrieveAttributes_WMS_AccessGroup, derefAliases_WMS_AccessGroup)
  # pp.pprint(wmsLinuxAccess_l)

  df_secures = pd.DataFrame({'WMS': wmsLinuxAccessGroupMembers_l})
  df_secures.to_excel(oshSecurityAuditWMS.wb2, sheet_name="WMS LDAP Access", index=False)

  oshSecurityAuditWMS.wb2.save()

  ## WMS - Linux groups:
  baseDN_WMS_LinuxGroups = "dc=orchard,dc=osh"
  searchScope_WMS_LinuxGroups = SUBTREE
  gidNumbers_WMS_LinuxGroups = ['10000', '10001', '10004', '10005']
  retrieveAttributes_WMS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  derefAliases_WMS_LinuxGroups = DEREF_ALWAYS

  gidMembers_d = oshSecurityAuditWMS.getLinuxGroupMembers(baseDN_WMS_LinuxGroups, searchScope_WMS_LinuxGroups,
                                                       retrieveAttributes_WMS_LinuxGroups, gidNumbers_WMS_LinuxGroups)
  WMS_Members_l = []
  for grpID in gidMembers_d.keys():

    # print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceWMS.DateTime, grpID))
    WMS_Members_l.append('{} - Linux Group ID [{}]:'.format(oshSecurityAuditWMS.DateTime, grpID))

    countGroupMembers = 0
    for member in gidMembers_d[grpID]:
      # print(member)
      WMS_Members_l.append(member)
      countGroupMembers += 1

    # print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
    WMS_Members_l.append('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))

  df_secures = pd.DataFrame({'WMS': WMS_Members_l})
  df_secures.to_excel(oshSecurityAuditWMS.wb2, sheet_name="WMS LDAP Groups", index=False)

  oshSecurityAuditWMS.wb2.save()
  #############################################################################################################


  # #############################################################################################################
  # ## MARS - Access Groups:
  # baseDN_MARS_LdapAccessGroups = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  # searchScope_MARS_LdapAccessGroup = SUBTREE
  # searchFilter_App_MARS_LdapAccessGroup = '(cn=Linux-MARS-App-Access*)'
  # searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  # retrieveAttributes_MARS_LdapAccessGroup = ['cn', 'member']
  # derefAliases_MARS_LdapAccessGroup = DEREF_ALWAYS
  #
  # pathMARS = r"C:/staging/python/pilot/testMARSAccountReview.xlsx"
  # stg_auditOSHSecurityComplianceMARS = osh_security_audit(ldapServer, userDN, passwd, pathMARS)
  #
  # marsLinuxAccess_l = stg_auditOSHSecurityComplianceMARS.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                 searchFilter_App_MARS_LdapAccessGroup,
  #                                                                                 retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup)
  # marsLinuxAccess_l.append(stg_auditOSHSecurityComplianceMARS.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                      searchFilter_DB_MARS_AccessGroup,
  #                                                                                      retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup))
  # # pp.pprint(marsLinuxAccess_l)
  # df_pollingLinuxAccess = pd.DataFrame({'MARS': marsLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Access", index=False)
  #
  # stg_auditOSHSecurityComplianceMARS.wb2.save()
  #
  # ## MARS - Linux groups:
  # baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  # searchScope_MARS_LinuxGroups = SUBTREE
  # gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  # retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  # derefAliases_MARS_LinuxGroups = DEREF_ALWAYS
  #
  # gidMembers_d = stg_auditOSHSecurityComplianceMARS.listLdapAdminGroupMembership(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
  #                                                      retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)
  #
  # for grpID in gidMembers_d.keys():
  #
  #   print('{} - Linux Group ID [{}]:'.format(stg_auditOSHSecurityComplianceMARS.DateTime, grpID))
  #
  #   countGroupMembers = 0
  #   for member in gidMembers_d[grpID]:
  #
  #     print(member)
  #     countGroupMembers += 1
  #
  #   print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #
  # df_pollingLinuxAccess = pd.DataFrame({'MARS': marsLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityComplianceMARS.wb2, sheet_name="MARS LDAP Groups", index=False)
  #
  # stg_auditOSHSecurityComplianceMARS.wb2.save()
  #
  # #############################################################################################################
  #
  #
  # #############################################################################################################
  # ## Polling - Access Groups:
  # baseDN_MARS_LdapAccessGroups = "ou=Groups,ou=Administrative,dc=orchard,dc=osh"
  # searchScope_MARS_LdapAccessGroup = SUBTREE
  # searchFilter_App_MARS_LdapAccessGroup = '(cn=Linux-MARS-App-Access*)'
  # searchFilter_DB_MARS_AccessGroup = '(cn=Linux-MARS-DB-Access*)'
  # retrieveAttributes_MARS_LdapAccessGroup = ['cn', 'member']
  # derefAliases_MARS_LdapAccessGroup = DEREF_ALWAYS
  #
  # pathPolling = r"C:/staging/python/pilot/testPollingAccountReview.xlsx"
  # stg_auditOSHSecurityCompliancePolling = osh_security_audit(ldapServer, userDN, passwd, pathPolling)
  #
  # pollingLinuxAccess_l = stg_auditOSHSecurityCompliancePolling.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                 searchFilter_App_MARS_LdapAccessGroup,
  #                                                                                 retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup)
  # pollingLinuxAccess_l.append(stg_auditOSHSecurityCompliancePolling.auditLinuxAdminLdapAccess(baseDN_MARS_LdapAccessGroups, searchScope_MARS_LdapAccessGroup,
  #                                                                                      searchFilter_DB_MARS_AccessGroup,
  #                                                                                      retrieveAttributes_MARS_LdapAccessGroup, derefAliases_MARS_LdapAccessGroup))
  # # pp.pprint(marsLinuxAccess_l)
  # df_pollingLinuxAccess = pd.DataFrame({'Polling': pollingLinuxAccess_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Access", index=False)
  #
  # stg_auditOSHSecurityCompliancePolling.wb2.save()
  #
  # ## MARS - Linux groups:
  # baseDN_MARS_LinuxGroups = "dc=orchard,dc=osh"
  # searchScope_MARS_LinuxGroups = SUBTREE
  # gidNumbers_MARS_LinuxGroups = ['10000', '10001', '10004', '10005']
  # retrieveAttributes_MARS_LinuxGroups = ['cn', 'gidNumber', 'uidNumber', 'loginShell', 'unixHomeDirectory']
  # derefAliases_MARS_LinuxGroups = DEREF_ALWAYS
  #
  # gidMembers_d = stg_auditOSHSecurityCompliancePolling.listLdapAdminGroupMembership(baseDN_MARS_LinuxGroups, searchScope_MARS_LinuxGroups,
  #                                                      retrieveAttributes_MARS_LinuxGroups, gidNumbers_MARS_LinuxGroups)
  #
  # pollingLdapAdminGroups_l = []
  # for grpID in gidMembers_d.keys():
  #
  #   print('{} - Linux Group ID [{}]:'.format(Enums.DateTime, grpID))
  #
  #   countGroupMembers = 0
  #   for member in gidMembers_d[grpID]:
  #
  #     # print(member)
  #     pollingLdapAdminGroups_l.append(member)
  #     countGroupMembers += 1
  #
  #   print('Total number of members in Linux Group [{}]: {}\n'.format(grpID, countGroupMembers))
  #
  #
  # df_pollingLinuxAccess = pd.DataFrame({'Polling': pollingLdapAdminGroups_l})
  # df_pollingLinuxAccess.to_excel(stg_auditOSHSecurityCompliancePolling.wb2, sheet_name="Polling LDAP Groups", index=False)
  #
  # stg_auditOSHSecurityCompliancePolling.wb2.save()
  # #############################################################################################################